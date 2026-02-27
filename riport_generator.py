"""
riport_generator.py
===================
Professzionális Excel riport generálása openpyxl-lel.

Javítások (v2):
  - ZeroDivisionError védelem _lap_temak-ban
  - Robustus None/NaN kezelés minden lapon
  - Timestamp a fájlnévben (felülírás elkerülése)
"""

from collections import Counter
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Stílus konstansok
# ---------------------------------------------------------------------------

_KEKFEJ         = "1F4E78"
_VILAGOSKEK     = "D9E1F2"
_SZURKE         = "F2F2F2"
_ZOLD_HATTER    = "C6EFCE"
_ZOLD_FONT      = "006100"
_PIROS_HATTER   = "FFC7CE"
_PIROS_FONT     = "9C0006"
_SARGA_HATTER   = "FFEB9C"
_SARGA_FONT     = "9C6500"

_VEKONY_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _fejlec_cella(cell, szoveg: str):
    """Kék fejléc cella formázás."""
    cell.value     = szoveg
    cell.font      = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill      = PatternFill(start_color=_KEKFEJ, end_color=_KEKFEJ, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = _VEKONY_BORDER


def _adat_cella(cell, ertek, alternalo: bool = False):
    """Adat cella alap formázás."""
    cell.value     = str(ertek) if ertek is not None else ""
    cell.font      = Font(name='Arial', size=10)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border    = _VEKONY_BORDER
    if alternalo:
        cell.fill = PatternFill(start_color=_SZURKE, fill_type='solid')


def _hibrid_szin(cell, kategoria: str):
    """Színkódolás a Végső Hibrid oszlophoz."""
    if kategoria == 'pozitív':
        cell.fill = PatternFill(start_color=_ZOLD_HATTER, fill_type='solid')
        cell.font = Font(name='Arial', size=10, bold=True, color=_ZOLD_FONT)
    elif kategoria == 'negatív':
        cell.fill = PatternFill(start_color=_PIROS_HATTER, fill_type='solid')
        cell.font = Font(name='Arial', size=10, bold=True, color=_PIROS_FONT)
    elif kategoria == 'semleges':
        cell.fill = PatternFill(start_color=_SARGA_HATTER, fill_type='solid')
        cell.font = Font(name='Arial', size=10, color=_SARGA_FONT)


# ---------------------------------------------------------------------------
# Lapok
# ---------------------------------------------------------------------------

def _lap_elemzesi_eredmenyek(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Elemzési Eredmények", 0)

    headers = [
        'Eredeti Szöveg', 'HuSpaCy Kategória', 'HunBERT Kategória',
        'Végső Hibrid', 'HunBERT Conf.', 'Téma', 'HuSpaCy Pont',
        'Pozitív Elemek', 'Negatív Elemek', 'Entitások',
        'Mondatok', 'Tokenek', 'Mondatszintű Bontás', 'Hibrid Megalapozás',
    ]
    oszlop_mapping = {
        'Eredeti Szöveg':       'eredeti_szoveg',
        'HuSpaCy Kategória':    'kategoria',
        'HunBERT Kategória':    'hunbert_kategoria',
        'Végső Hibrid':         'hibrid_kategoria',
        'HunBERT Conf.':        'hunbert_confidence',
        'Téma':                 'tema_kulcsszavak',
        'HuSpaCy Pont':         'pontszam',
        'Pozitív Elemek':       'pozitiv_elemek',
        'Negatív Elemek':       'negativ_elemek',
        'Entitások':            'emlitett_nevek',
        'Mondatok':             'mondatok_szama',
        'Tokenek':              'token_szam',
        'Mondatszintű Bontás':  'mondatszintu_sentiment',
        'Hibrid Megalapozás':   'hibrid_megalapozas',
    }

    for c_idx, h in enumerate(headers, 1):
        _fejlec_cella(ws.cell(row=1, column=c_idx), h)

    ws.row_dimensions[1].height = 35

    records = df.to_dict('records')
    for r_idx, row_dict in enumerate(records, 2):
        alternalo = (r_idx % 2 == 0)
        for c_idx, h in enumerate(headers, 1):
            col_name = oszlop_mapping.get(h, h.lower().replace(' ', '_'))
            val      = row_dict.get(col_name, '')
            # NaN / None kezelés
            if pd.isna(val) if not isinstance(val, (str, list, dict)) else False:
                val = ''
            cell = ws.cell(row=r_idx, column=c_idx)
            _adat_cella(cell, val, alternalo)
            if h == 'Végső Hibrid':
                _hibrid_szin(cell, str(val))

    # Oszlopszélességek
    szelessegek = [50, 15, 15, 15, 12, 30, 12, 25, 25, 25, 10, 10, 40, 50]
    for i, w in enumerate(szelessegek, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sor_max = max(len(df) + 1, 2)

    # Feltételes formázás – HuSpaCy pont
    ws.conditional_formatting.add(
        f"G2:G{sor_max}",
        DataBarRule(
            start_type='num', start_value=-1,
            end_type='num', end_value=1,
            color='638EC6',
        ),
    )
    # Feltételes formázás – HunBERT konfidencia
    ws.conditional_formatting.add(
        f"E2:E{sor_max}",
        ColorScaleRule(
            start_type='num', start_value=0,   start_color='F8696B',
            mid_type='num',   mid_value=0.5,   mid_color='FFEB84',
            end_type='num',   end_value=1,     end_color='63BE7B',
        ),
    )


def _lap_statisztika(wb: Workbook, df: pd.DataFrame, stats: dict):
    ws = wb.create_sheet("Statisztikai Összefoglaló")

    ws['A1'] = 'HIBRID SENTIMENT ELEMZÉS – ÖSSZEFOGLALÓ'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color=_KEKFEJ)
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30

    ws['A3'] = 'Kategória Eloszlás'
    ws['A3'].font = Font(name='Arial', size=12, bold=True)

    for col, txt in zip(['A', 'B', 'C'], ['Kategória', 'Darabszám', 'Százalék']):
        cell       = ws[f'{col}4']
        cell.value = txt
        cell.font  = Font(name='Arial', bold=True)
        cell.fill  = PatternFill(start_color=_VILAGOSKEK, fill_type='solid')
        cell.border = _VEKONY_BORDER

    kategoriak = ['Pozitív (Hibrid)', 'Negatív (Hibrid)', 'Semleges (Hibrid)']
    szin_map = {
        'Pozitív (Hibrid)':  _ZOLD_HATTER,
        'Negatív (Hibrid)':  _PIROS_HATTER,
        'Semleges (Hibrid)': _SARGA_HATTER,
    }

    total = sum(stats.get(k, 0) for k in kategoriak)

    row = 5
    for kat in kategoriak:
        count = stats.get(kat, 0)
        label = kat.replace(' (Hibrid)', '')

        a_cell         = ws[f'A{row}']
        a_cell.value   = label
        a_cell.fill    = PatternFill(start_color=szin_map[kat], fill_type='solid')
        a_cell.border  = _VEKONY_BORDER

        b_cell         = ws[f'B{row}']
        b_cell.value   = count
        b_cell.border  = _VEKONY_BORDER

        pct_cell           = ws[f'C{row}']
        pct_cell.value     = (count / total * 100) if total > 0 else 0.0
        pct_cell.number_format = '0.0"%"'
        pct_cell.border    = _VEKONY_BORDER

        row += 1

    # Átlagos pontszámok
    ws[f'A{row + 1}'] = 'Átlagos HuSpaCy Pontszám'
    ws[f'A{row + 1}'].font = Font(name='Arial', bold=True)
    avg_cell               = ws[f'B{row + 1}']
    avg_cell.value         = float(stats.get('Átlagos HuSpaCy pont', 0))
    avg_cell.number_format = '0.000'

    ws[f'A{row + 2}'] = 'Átlagos HunBERT Konfidencia'
    ws[f'A{row + 2}'].font = Font(name='Arial', bold=True)
    if 'hunbert_confidence' in df.columns:
        conf_cell               = ws[f'B{row + 2}']
        try:
            conf_cell.value = float(df['hunbert_confidence'].mean())
        except Exception:
            conf_cell.value = 0.0
        conf_cell.number_format = '0.0%'

    ws[f'A{row + 3}'] = 'Összes elemzett szöveg'
    ws[f'A{row + 3}'].font = Font(name='Arial', bold=True)
    ws[f'B{row + 3}'] = len(df)

    # Pie chart
    pie         = PieChart()
    pie.title   = "Sentiment Kategória Eloszlás"
    pie.style   = 10
    pie.width   = 12
    pie.height  = 10
    data        = Reference(ws, min_col=2, min_row=4, max_row=7)
    labels      = Reference(ws, min_col=1, min_row=5, max_row=7)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, 'E3')

    # Bar chart
    bar             = BarChart()
    bar.title       = "Kategóriák Összehasonlítása"
    bar.style       = 11
    bar.width       = 15
    bar.height      = 10
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.y_axis.title = 'Darabszám'
    ws.add_chart(bar, 'E20')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12


def _lap_temak(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Témaelemzés")
    ws['A1'] = 'BERTOPIC TÉMAKÖR ELEMZÉS'
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color=_KEKFEJ)
    ws.merge_cells('A1:C1')

    for col, txt in zip(['A', 'B', 'C'], ['Téma', 'Előfordulás', 'Százalék']):
        cell        = ws[f'{col}3']
        cell.value  = txt
        cell.font   = Font(name='Arial', bold=True)
        cell.fill   = PatternFill(start_color=_VILAGOSKEK, fill_type='solid')
        cell.border = _VEKONY_BORDER

    if 'tema_kulcsszavak' in df.columns:
        tema_counts = df['tema_kulcsszavak'].value_counts()
        total       = len(df)

        # ZeroDivisionError védelem: total > 0 garantált ha van adat
        for row, (tema, count) in enumerate(tema_counts.head(15).items(), 4):
            a_cell        = ws[f'A{row}']
            a_cell.value  = str(tema)[:50]
            a_cell.border = _VEKONY_BORDER

            b_cell        = ws[f'B{row}']
            b_cell.value  = int(count)
            b_cell.border = _VEKONY_BORDER

            pct_cell            = ws[f'C{row}']
            pct_cell.value      = (count / total * 100) if total > 0 else 0.0
            pct_cell.number_format = '0.0"%"'
            pct_cell.border     = _VEKONY_BORDER

            if row % 2 == 0:
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color=_SZURKE, fill_type='solid')

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12


def _lap_nyelveszet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Nyelvészeti Részletek")
    ws['A1'] = 'HUSPACY NYELVÉSZETI MÉLYSTRUKTÚRA'
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color=_KEKFEJ)
    ws.merge_cells('A1:F1')

    fejlecek = [
        'Szövegrészlet', 'Pozitív Elemek', 'Negatív Elemek',
        'Szintaktikai Kapcsolatok', 'POS Stat', 'DEP Stat',
    ]
    oszlopok = [
        'eredeti_szoveg', 'pozitiv_elemek', 'negativ_elemek',
        'nyelveszeti_kapcsolatok', 'pos_statisztika', 'dep_statisztika',
    ]

    for c_idx, h in enumerate(fejlecek, 1):
        cell           = ws.cell(row=3, column=c_idx, value=h)
        cell.font      = Font(name='Arial', bold=True)
        cell.fill      = PatternFill(start_color=_VILAGOSKEK, fill_type='solid')
        cell.alignment = Alignment(wrap_text=True)
        cell.border    = _VEKONY_BORDER

    for r_idx, row_dict in enumerate(df.to_dict('records')[:100], 4):
        alternalo = (r_idx % 2 == 0)
        for c_idx, col_name in enumerate(oszlopok, 1):
            val = row_dict.get(col_name, '')
            if pd.isna(val) if not isinstance(val, (str, list, dict)) else False:
                val = ''
            if col_name == 'eredeti_szoveg':
                val = str(val)[:100] + ('...' if len(str(val)) > 100 else '')
            _adat_cella(ws.cell(row=r_idx, column=c_idx), val, alternalo)

    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F'], [40, 30, 30, 50, 25, 25]):
        ws.column_dimensions[col].width = w


def _lap_entitasok(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Entitások")
    ws['A1'] = 'NAMED ENTITY RECOGNITION (NER)'
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color=_KEKFEJ)
    ws.merge_cells('A1:C1')

    for col, txt in zip(['A', 'B', 'C'], ['Entitás', 'Előfordulás', 'Típus']):
        cell        = ws[f'{col}3']
        cell.value  = txt
        cell.font   = Font(name='Arial', bold=True)
        cell.fill   = PatternFill(start_color=_VILAGOSKEK, fill_type='solid')
        cell.border = _VEKONY_BORDER

    if 'emlitett_nevek' in df.columns:
        all_entities = []
        for ents_str in df['emlitett_nevek'].dropna():
            s = str(ents_str)
            if s:
                for ent in s.split(', '):
                    if '(' in ent and ')' in ent:
                        all_entities.append(ent)

        ent_counts = Counter(all_entities)
        for row, (ent, count) in enumerate(ent_counts.most_common(30), 4):
            parts = ent.split('(')
            name  = parts[0].strip()
            tipo  = parts[1].rstrip(')') if len(parts) > 1 else 'N/A'

            a_cell        = ws[f'A{row}']
            a_cell.value  = name
            a_cell.border = _VEKONY_BORDER

            b_cell        = ws[f'B{row}']
            b_cell.value  = count
            b_cell.border = _VEKONY_BORDER

            c_cell        = ws[f'C{row}']
            c_cell.value  = tipo
            c_cell.border = _VEKONY_BORDER

            if row % 2 == 0:
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color=_SZURKE, fill_type='solid')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


# ---------------------------------------------------------------------------
# Fő függvény
# ---------------------------------------------------------------------------

def generalj_riportot(
    df: pd.DataFrame,
    stats: dict,
    kimenet: str = None,
) -> str:
    """
    Professzionális Excel riport generálása teljes formázással.

    Ha kimenet=None, automatikus timestampes fájlnevet generál
    (elkerüli a korábbi riportok felülírását).

    Visszatér: a létrehozott fájl neve.
    """
    if kimenet is None:
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        kimenet = f"sentiment_riport_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # Alapértelmezett üres lap eltávolítása

    _lap_elemzesi_eredmenyek(wb, df)
    _lap_statisztika(wb, df, stats)
    _lap_temak(wb, df)
    _lap_nyelveszet(wb, df)
    _lap_entitasok(wb, df)

    wb.save(kimenet)
    print(f"Professzionális Excel riport elkészült: {kimenet}")
    return kimenet
